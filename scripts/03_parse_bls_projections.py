#!/usr/bin/env python3
"""
Step 3: Download and parse BLS Employment Projections 2024-2034.

Downloads the BLS occupation projections workbook and extracts Table 1.2
(Occupational projections, 2024-2034, and worker characteristics).

Extracts per occupation:
  - 2024 employment (thousands)
  - 2034 projected employment (thousands)
  - Employment change (numeric and percent)
  - Computed growth rate
  - Median annual wage (where available)
  - Education/training requirements

Then harmonizes SOC codes with our O*NET-derived skill profiles (774 occupations)
and produces a coverage report.

Outputs:
  - data/bls_projections.json

Data source: BLS Employment Projections, published 2025
  https://www.bls.gov/emp/ind-occ-matrix/occupation.xlsx
  Table 1.2: Occupational projections, 2024-2034

Requires: openpyxl (pip install openpyxl)
"""

import json
import os
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

BLS_URL = "https://www.bls.gov/emp/ind-occ-matrix/occupation.xlsx"
BLS_LOCAL = Path("/tmp/bls_occupation_matrix.xlsx")

ONET_PROFILES_FILE = DATA_DIR / "onet_skill_profiles.json"
OUTPUT_FILE = DATA_DIR / "bls_projections.json"


def download_bls_file():
    """Download BLS occupation matrix Excel file if not already present."""
    if BLS_LOCAL.exists():
        size = BLS_LOCAL.stat().st_size
        if size > 100_000:  # expect ~400KB; skip if already downloaded
            print(f"  Using cached file: {BLS_LOCAL} ({size / 1024:.0f} KB)")
            return

    print(f"  Downloading from {BLS_URL}...")
    result = subprocess.run(
        [
            "curl", "-L", "-o", str(BLS_LOCAL), BLS_URL,
            "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "-H", f"Referer: https://www.bls.gov/emp/data/occupational-data.htm",
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"  ERROR: curl failed: {result.stderr}")
        sys.exit(1)

    size = BLS_LOCAL.stat().st_size
    print(f"  Downloaded {size / 1024:.0f} KB")

    # Verify it's actually an Excel file
    with open(BLS_LOCAL, "rb") as f:
        magic = f.read(4)
    if magic[:2] != b"PK":  # xlsx is a zip file
        print("  ERROR: Downloaded file is not a valid xlsx (may be a redirect/block page)")
        print("  Try downloading manually from:")
        print(f"    {BLS_URL}")
        print(f"  Save to: {BLS_LOCAL}")
        sys.exit(1)


def parse_wage(value):
    """Parse wage value, handling special cases like '—' and '>=$239,200'."""
    if value is None or value == "—" or value == "–":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        # Handle ">=$239,200" top-coded wages
        cleaned = value.replace(",", "").replace("$", "").replace(">=", "").strip()
        try:
            return int(float(cleaned))
        except (ValueError, TypeError):
            return None
    return None


def parse_numeric(value):
    """Parse a numeric value, returning None for missing/special values."""
    if value is None or value == "—" or value == "–":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    return None


def parse_table_1_2(filepath):
    """
    Parse Table 1.2 from the BLS occupation matrix workbook.

    Returns list of dicts with occupation-level projections data.
    Only returns "Line item" rows (detailed occupations), not summaries.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table 1.2"]

    occupations = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        # Column mapping (0-indexed):
        # 0: title, 1: SOC code, 2: occupation type, 3: emp 2024, 4: emp 2034
        # 5: emp dist 2024, 6: emp dist 2034, 7: emp change numeric
        # 8: emp change pct, 9: pct self-employed, 10: annual openings
        # 11: median wage, 12: education, 13: work experience, 14: training

        occ_type = row[2]
        soc_code = row[1]

        # Skip non-line-item rows (summaries, footnotes, blanks)
        if occ_type != "Line item":
            continue

        # Skip rows without valid SOC codes
        if not soc_code or not isinstance(soc_code, str):
            continue

        soc_code = soc_code.strip()

        # Validate SOC code format (XX-XXXX)
        if len(soc_code) != 7 or soc_code[2] != "-":
            continue

        title = row[0].strip() if row[0] else soc_code
        emp_2024 = parse_numeric(row[3])
        emp_2034 = parse_numeric(row[4])
        emp_change = parse_numeric(row[7])
        emp_change_pct = parse_numeric(row[8])
        annual_openings = parse_numeric(row[10])
        median_wage = parse_wage(row[11])
        education = row[12] if row[12] and row[12] != "—" else None
        work_experience = row[13] if row[13] and row[13] != "—" else None
        training = row[14] if row[14] and row[14] != "—" else None

        # Compute growth rate
        if emp_2024 and emp_2024 > 0 and emp_2034 is not None:
            growth_rate = round((emp_2034 - emp_2024) / emp_2024, 6)
        else:
            growth_rate = None

        occupations.append({
            "soc_code": soc_code,
            "title": title,
            "employment_2024": emp_2024,
            "employment_2034": emp_2034,
            "employment_change": emp_change,
            "employment_change_pct": emp_change_pct,
            "growth_rate": growth_rate,
            "annual_openings": annual_openings,
            "median_annual_wage": median_wage,
            "typical_education": education,
            "work_experience": work_experience,
            "typical_training": training,
        })

    return occupations


def load_onet_soc_codes(filepath):
    """Load SOC codes from O*NET skill profiles for coverage comparison."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return set(data["profiles"].keys())


def run_sanity_checks(occupations, onet_codes, matched, ep_only, onet_only):
    """Run comprehensive sanity checks on the parsed data."""
    print("\n" + "=" * 60)
    print("SANITY CHECKS")
    print("=" * 60)

    all_pass = True

    def check(name, condition, detail=""):
        nonlocal all_pass
        status = "PASS" if condition else "FAIL"
        if not condition:
            all_pass = False
        print(f"  [{status}] {name}" + (f" — {detail}" if detail else ""))

    # 1. Row count
    check(
        "Occupation count in expected range (700-900)",
        700 <= len(occupations) <= 900,
        f"{len(occupations)} line-item occupations",
    )

    # 2. Total employment
    total_emp = sum(o["employment_2024"] for o in occupations if o["employment_2024"])
    check(
        "Total 2024 employment ~170M (150M-190M range)",
        150_000 < total_emp < 190_000,  # in thousands
        f"{total_emp:.1f} thousand = {total_emp / 1000:.1f} million",
    )

    # 3. No duplicate SOC codes
    soc_codes = [o["soc_code"] for o in occupations]
    unique = len(set(soc_codes))
    check("No duplicate SOC codes", unique == len(soc_codes), f"{unique} unique / {len(soc_codes)} total")

    # 4. No null employment values
    null_emp = sum(1 for o in occupations if o["employment_2024"] is None or o["employment_2034"] is None)
    check("No null employment values", null_emp == 0, f"{null_emp} occupations with null employment")

    # 5. Growth rate distribution
    growth_rates = [o["growth_rate"] for o in occupations if o["growth_rate"] is not None]
    extreme = [g for g in growth_rates if g < -0.50 or g > 1.0]
    check(
        "No extreme growth rates (all between -50% and +100%)",
        len(extreme) == 0,
        f"{len(extreme)} extreme values" if extreme else "all within range",
    )

    import numpy as np
    gr = np.array(growth_rates)
    print(f"\n  Growth rate distribution:")
    print(f"    Min: {gr.min():.4f}  Max: {gr.max():.4f}")
    print(f"    Mean: {gr.mean():.4f}  Median: {np.median(gr):.4f}")
    for pct in [5, 25, 50, 75, 95]:
        print(f"    {pct}th percentile: {np.percentile(gr, pct):.4f}")

    # 6. Known-value spot checks
    print(f"\n  Known-value spot checks:")
    occ_lookup = {o["soc_code"]: o for o in occupations}

    spot_checks = {
        "15-1252": ("Software Developers", 1500, 2200),  # rough expected range
        "29-1141": ("Registered Nurses", 3000, 4000),
        "11-1021": ("General and Operations Managers", 3000, 4500),
    }

    for soc, (expected_title, min_emp, max_emp) in spot_checks.items():
        if soc in occ_lookup:
            o = occ_lookup[soc]
            emp = o["employment_2024"]
            in_range = min_emp < emp < max_emp
            check(
                f"{soc} ({o['title']}): employment in range",
                in_range,
                f"employment={emp:.1f}K (expected {min_emp}-{max_emp}K)",
            )
        else:
            check(f"{soc} ({expected_title}) found in data", False, "NOT FOUND")

    # 7. Coverage
    print(f"\n  Coverage report:")
    check(
        "High O*NET match rate (>90%)",
        len(matched) / len(onet_codes) > 0.90,
        f"{len(matched)}/{len(onet_codes)} = {len(matched)/len(onet_codes)*100:.1f}%",
    )
    print(f"    EP-only (no O*NET profile): {len(ep_only)} occupations")
    print(f"    O*NET-only (no EP data): {len(onet_only)} occupations")

    # 8. Wage coverage
    wage_count = sum(1 for o in occupations if o["median_annual_wage"] is not None)
    check(
        "Wage data available for >95% of occupations",
        wage_count / len(occupations) > 0.95,
        f"{wage_count}/{len(occupations)} = {wage_count/len(occupations)*100:.1f}%",
    )

    print(f"\n  Overall: {'ALL CHECKS PASSED' if all_pass else 'SOME CHECKS FAILED'}")
    return all_pass


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("Step 3: Download and Parse BLS Employment Projections")
    print("=" * 60)

    # Download
    print("\nDownloading BLS data...")
    download_bls_file()

    # Parse
    print("\nParsing Table 1.2 (Occupational projections, 2024-2034)...")
    occupations = parse_table_1_2(BLS_LOCAL)
    print(f"  Parsed {len(occupations)} detailed occupations")

    # Load O*NET codes for coverage check
    print("\nComparing with O*NET skill profiles...")
    if ONET_PROFILES_FILE.exists():
        onet_codes = load_onet_soc_codes(ONET_PROFILES_FILE)
        print(f"  O*NET profiles: {len(onet_codes)} occupations")
    else:
        print(f"  WARNING: {ONET_PROFILES_FILE} not found. Skipping coverage check.")
        onet_codes = set()

    ep_codes = set(o["soc_code"] for o in occupations)
    matched = onet_codes & ep_codes
    ep_only = ep_codes - onet_codes
    onet_only = onet_codes - ep_codes

    print(f"  Matched: {len(matched)}")
    print(f"  EP-only (no O*NET profile): {len(ep_only)}")
    print(f"  O*NET-only (no EP data): {len(onet_only)}")

    # Sanity checks
    all_pass = run_sanity_checks(occupations, onet_codes, matched, ep_only, onet_only)

    # Build output
    print("\n" + "=" * 60)
    print("WRITING OUTPUT")
    print("=" * 60)

    output = {
        "metadata": {
            "source": "BLS Employment Projections 2024-2034",
            "source_table": "Table 1.2: Occupational projections, 2024-2034, and worker characteristics",
            "url": BLS_URL,
            "employment_unit": "thousands",
            "base_year": 2024,
            "projection_year": 2034,
            "total_occupations": len(occupations),
            "matched_to_onet_profiles": len(matched),
            "ep_only_no_onet_profile": len(ep_only),
            "onet_only_no_ep_data": len(onet_only),
            "wage_note": "Wages from OEWS May 2024. Values of 239200 indicate top-coded (>=$239,200). Null indicates unavailable.",
            "generated_by": "scripts/03_parse_bls_projections.py",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "projections": {},
        "coverage": {
            "matched": sorted(matched),
            "ep_only": sorted(ep_only),
            "onet_only": sorted(onet_only),
        },
    }

    for occ in sorted(occupations, key=lambda x: x["soc_code"]):
        soc = occ["soc_code"]
        output["projections"][soc] = {
            "title": occ["title"],
            "employment_2024": occ["employment_2024"],
            "employment_2034": occ["employment_2034"],
            "employment_change": occ["employment_change"],
            "employment_change_pct": occ["employment_change_pct"],
            "growth_rate": occ["growth_rate"],
            "annual_openings": occ["annual_openings"],
            "median_annual_wage": occ["median_annual_wage"],
            "typical_education": occ["typical_education"],
            "work_experience": occ["work_experience"],
            "typical_training": occ["typical_training"],
        }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)
    file_size = OUTPUT_FILE.stat().st_size / (1024 * 1024)
    print(f"\n  Wrote {OUTPUT_FILE} ({file_size:.2f} MB)")

    # Sample output
    print("\n" + "=" * 60)
    print("SAMPLE OUTPUT (first 3 occupations)")
    print("=" * 60)

    sample_codes = sorted(output["projections"].keys())[:3]
    for soc in sample_codes:
        p = output["projections"][soc]
        print(f"\n{soc}: {p['title']}")
        print(f"  Employment 2024: {p['employment_2024']:.1f}K ({p['employment_2024']*1000:,.0f} workers)")
        print(f"  Employment 2034: {p['employment_2034']:.1f}K")
        print(f"  Growth rate: {p['growth_rate']:.4f} ({p['employment_change_pct']}%)")
        print(f"  Annual openings: {p['annual_openings']:.1f}K")
        print(f"  Median wage: ${p['median_annual_wage']:,}" if p["median_annual_wage"] else "  Median wage: N/A")
        print(f"  Education: {p['typical_education']}")

    # Print some O*NET-only codes for investigation
    if onet_only:
        print("\n" + "=" * 60)
        print(f"O*NET-ONLY CODES ({len(onet_only)} occupations with no EP data)")
        print("=" * 60)
        # Load O*NET titles for context
        if ONET_PROFILES_FILE.exists():
            with open(ONET_PROFILES_FILE) as f:
                onet_data = json.load(f)
            for soc in sorted(onet_only):
                title = onet_data["profiles"].get(soc, {}).get("title", "Unknown")
                print(f"  {soc}: {title}")


if __name__ == "__main__":
    main()
