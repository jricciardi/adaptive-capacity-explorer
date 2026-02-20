#!/usr/bin/env python3
"""
Step 5: Compute geographic (metro) density scores.

For each occupation, computes:
  E[log(density)|occupation] = Σi( emp_occ_i / emp_occ_total ) × log( emp_all_i / area_i )

This is the employment-weighted average of log overall CBSA density across
metros where an occupation's workers are located. Higher scores mean the
occupation's workers are concentrated in denser (more jobs per sq mile) metros.

Data sources:
  - BLS OEWS May 2024: per-occupation and total metro employment
  - Census 2024 CBSA Gazetteer: land area per CBSA

Inputs:
  - data/transferability_scores.json (from Step 4, for 751 matched SOC codes)

Outputs:
  - data/density_scores.json

Based on Manning & Aguirre (2026), "How Adaptable Are American Workers
to AI-Induced Job Displacement?", NBER Working Paper 34705.

Requires: openpyxl (pip install openpyxl), numpy
"""

import csv
import json
import math
import subprocess
import sys
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import numpy as np

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

TRANSFERABILITY_FILE = DATA_DIR / "transferability_scores.json"
OUTPUT_FILE = DATA_DIR / "density_scores.json"

# Download URLs and cache paths
OEWS_URL = "https://www.bls.gov/oes/special.requests/oesm24ma.zip"
OEWS_ZIP = Path("/tmp/oews_msa_may2024.zip")
OEWS_DIR = Path("/tmp/oews_msa_may2024")

GAZETTEER_URL = (
    "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/"
    "2024_Gazetteer/2024_Gaz_cbsa_national.zip"
)
GAZETTEER_ZIP = Path("/tmp/census_gazetteer_cbsa_2024.zip")
GAZETTEER_DIR = Path("/tmp/census_gazetteer_cbsa_2024")

# User-Agent header for BLS downloads (BLS blocks default requests)
CURL_HEADERS = [
    "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "-H", "Referer: https://www.bls.gov/oes/tables.htm",
]


def download_file(url, local_path, min_size, description):
    """Download a file via curl with caching and validation.

    BLS actively blocks automated downloads (Akamai bot detection).
    If the download fails or returns HTML instead of a ZIP, the script
    provides manual download instructions.
    """
    if local_path.exists():
        size = local_path.stat().st_size
        if size > min_size:
            # Verify it's actually a ZIP and not a cached HTML error page
            with open(local_path, "rb") as f:
                magic = f.read(4)
            if magic[:2] == b"PK":
                print(f"  Using cached file: {local_path} ({size / 1024:.0f} KB)")
                return
            else:
                print(f"  Cached file is invalid (HTML/error page), re-downloading...")
                local_path.unlink()

    print(f"  Downloading {description}...")
    print(f"    URL: {url}")
    result = subprocess.run(
        ["curl", "-L", "-o", str(local_path), url] + CURL_HEADERS,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"  ERROR: curl failed: {result.stderr}")
        _print_manual_download_instructions(url, local_path)
        sys.exit(1)

    size = local_path.stat().st_size
    print(f"  Downloaded {size / 1024:.0f} KB")

    # Verify it's a valid ZIP
    with open(local_path, "rb") as f:
        magic = f.read(4)
    if magic[:2] != b"PK":
        local_path.unlink()  # Remove invalid file
        _print_manual_download_instructions(url, local_path)
        sys.exit(1)


def _print_manual_download_instructions(url, local_path):
    """Print instructions for manual download when BLS blocks automated access."""
    print()
    print("  " + "=" * 56)
    print("  BLS blocked the automated download (bot detection).")
    print("  Please download manually:")
    print()
    print(f"  1. Open in your browser: {url}")
    print(f"  2. Save the file to: {local_path}")
    print()
    print("  Then re-run this script.")
    print("  " + "=" * 56)


def extract_zip(zip_path, extract_dir):
    """Extract a ZIP file to a directory."""
    if extract_dir.exists() and any(extract_dir.iterdir()):
        print(f"  Using cached extraction: {extract_dir}")
        return

    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_dir)
    files = list(extract_dir.iterdir())
    print(f"  Extracted {len(files)} file(s) to {extract_dir}")
    for f in files:
        print(f"    {f.name} ({f.stat().st_size / 1024:.0f} KB)")


def parse_gazetteer(gazetteer_dir):
    """
    Parse Census CBSA gazetteer for land area by CBSA.

    Returns:
        cbsa_areas: dict of cbsa_code (str) -> {name, land_area_sq_mi}
    """
    # Find the text file
    txt_files = list(gazetteer_dir.glob("*.txt"))
    if not txt_files:
        print("  ERROR: No .txt file found in gazetteer directory")
        sys.exit(1)

    gazetteer_file = txt_files[0]
    print(f"  Parsing {gazetteer_file.name}...")

    cbsa_areas = {}
    with open(gazetteer_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        # Strip whitespace from field names (Census files often have trailing spaces)
        reader.fieldnames = [name.strip() for name in reader.fieldnames]

        for row in reader:
            # Strip all values too
            row = {k.strip(): v.strip() if isinstance(v, str) else v for k, v in row.items()}

            geoid = row.get("GEOID", "").strip()
            name = row.get("NAME", "").strip()
            aland_sqmi = row.get("ALAND_SQMI", "").strip()

            if not geoid or not aland_sqmi:
                continue

            try:
                area = float(aland_sqmi)
            except (ValueError, TypeError):
                continue

            if area <= 0:
                continue

            cbsa_areas[geoid] = {
                "name": name,
                "land_area_sq_mi": area,
            }

    return cbsa_areas


def parse_employment(value):
    """Parse OEWS TOT_EMP value, returning None for suppressed values."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    s = str(value).strip()
    if s in ("**", "*", "~", "#", "", "N/A", "N", "na"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def parse_oews_msa(oews_dir):
    """
    Parse OEWS MSA flat file for per-metro occupation employment.

    Filters:
    - AREA_TYPE = 2 (Metropolitan Statistical Areas)
    - Cross-industry totals (NAICS starts with "000000" or similar)
    - Detailed occupations (O_GROUP = "detailed") + All Occupations (00-0000)

    Returns:
        occ_by_metro: dict of soc_code -> {cbsa: employment}
        metro_total_emp: dict of cbsa -> total employment (from 00-0000 row)
        metro_names: dict of cbsa -> area title
    """
    # Find the MSA XLSX file (may be in a subdirectory within the ZIP)
    # Look for files matching MSA_*.xlsx or oesm*.xlsx pattern
    xlsx_files = list(oews_dir.glob("**/*.xlsx"))
    # Filter to the MSA data file (not file_descriptions.xlsx, not BOS_*.xlsx)
    msa_files = [f for f in xlsx_files if "MSA" in f.name.upper() or "oesm" in f.name.lower()]
    if not msa_files:
        # Fall back to any xlsx that isn't a description file
        msa_files = [f for f in xlsx_files if "description" not in f.name.lower()]
    if not msa_files:
        # Try CSV as fallback
        csv_files = list(oews_dir.glob("**/*.csv"))
        if csv_files:
            return _parse_oews_csv(csv_files[0])
        print("  ERROR: No OEWS MSA data file found in directory")
        print(f"  Files found: {[f.name for f in xlsx_files]}")
        sys.exit(1)

    xlsx_file = msa_files[0]
    print(f"  Parsing {xlsx_file.name} (this may take 30-60 seconds)...")

    wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in header_row]
    col_idx = {name: i for i, name in enumerate(headers)}

    # Verify required columns exist
    required = ["AREA", "AREA_TYPE", "OCC_CODE", "O_GROUP", "TOT_EMP", "AREA_TITLE"]
    missing = [c for c in required if c not in col_idx]
    if missing:
        # Try case-insensitive match
        header_upper = {h.upper(): h for h in headers}
        for m in missing:
            if m.upper() in header_upper:
                col_idx[m] = col_idx[header_upper[m.upper()]]
        missing = [c for c in required if c not in col_idx]
        if missing:
            print(f"  ERROR: Missing columns in OEWS file: {missing}")
            print(f"  Available columns: {headers[:20]}...")
            sys.exit(1)

    occ_by_metro = defaultdict(dict)
    metro_total_emp = {}
    metro_names = {}

    rows_read = 0
    rows_used = 0
    area_types_seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_read += 1

        area = str(row[col_idx["AREA"]]).strip()
        area_type = str(row[col_idx["AREA_TYPE"]]).strip()
        occ_code = str(row[col_idx["OCC_CODE"]]).strip()
        o_group = str(row[col_idx["O_GROUP"]]).strip()
        tot_emp_raw = row[col_idx["TOT_EMP"]]
        area_title = str(row[col_idx["AREA_TITLE"]]).strip()

        area_types_seen.add(area_type)

        # Only Metropolitan Statistical Areas
        # OEWS MSA flat file uses AREA_TYPE=4 for MSAs (not 2 as in the
        # combined all-areas file). Accept both for robustness.
        if area_type not in ("2", "4"):
            continue

        # Only cross-industry totals
        if "NAICS" in col_idx:
            naics = str(row[col_idx["NAICS"]]).strip()
            if naics not in ("000000", "000000 "):
                continue

        # Parse employment
        emp = parse_employment(tot_emp_raw)
        if emp is None:
            continue

        metro_names[area] = area_title
        rows_used += 1

        # All Occupations row -> total metro employment
        if occ_code == "00-0000":
            metro_total_emp[area] = emp
            continue

        # Only detailed occupations for per-occupation data
        if o_group != "detailed":
            continue

        occ_by_metro[occ_code][area] = emp

    wb.close()

    print(f"  Read {rows_read:,} rows, used {rows_used:,} MSA rows")
    print(f"  Area types seen: {sorted(area_types_seen)}")

    return dict(occ_by_metro), metro_total_emp, metro_names


def compute_cbsa_log_density(metro_total_emp, cbsa_areas):
    """
    Compute log(employment density) for each CBSA.

    density = total_employment / land_area_sq_mi
    log_density = ln(density)

    Returns:
        cbsa_log_density: dict of cbsa -> log_density value
        cbsa_info: dict of cbsa -> {name, total_employment, land_area_sq_mi, log_density}
    """
    cbsa_log_density = {}
    cbsa_info = {}
    skipped = 0

    for cbsa, total_emp in metro_total_emp.items():
        if cbsa not in cbsa_areas:
            skipped += 1
            continue

        area_data = cbsa_areas[cbsa]
        area_sq_mi = area_data["land_area_sq_mi"]

        if total_emp <= 0 or area_sq_mi <= 0:
            skipped += 1
            continue

        density = total_emp / area_sq_mi
        log_d = math.log(density)

        cbsa_log_density[cbsa] = log_d
        cbsa_info[cbsa] = {
            "name": area_data["name"],
            "total_employment": int(total_emp),
            "land_area_sq_mi": round(area_sq_mi, 1),
            "log_density": round(log_d, 4),
        }

    if skipped > 0:
        print(f"  Skipped {skipped} CBSAs (no gazetteer match or invalid data)")

    return cbsa_log_density, cbsa_info


def compute_density_scores(occ_by_metro, cbsa_log_density, matched_codes, metro_names):
    """
    Compute E[log(density)|occupation] for each occupation.

    For each occupation, computes the employment-weighted average of
    log(density) across the metros where that occupation has workers.

    Args:
        occ_by_metro: {soc: {cbsa: employment}}
        cbsa_log_density: {cbsa: log_density}
        matched_codes: list of target SOC codes
        metro_names: {cbsa: name}

    Returns:
        scores: dict of soc -> {density_score, cbsa_count, national_employment, top_metros}
        unscored: list of SOC codes that couldn't be scored
    """
    scores = {}
    unscored = []

    for soc in matched_codes:
        if soc not in occ_by_metro:
            unscored.append(soc)
            continue

        metro_emp = occ_by_metro[soc]

        # Filter to CBSAs where we have log-density
        valid_metros = {
            cbsa: emp for cbsa, emp in metro_emp.items()
            if cbsa in cbsa_log_density
        }

        if not valid_metros:
            unscored.append(soc)
            continue

        # National total across valid metros
        emp_total = sum(valid_metros.values())
        if emp_total <= 0:
            unscored.append(soc)
            continue

        # Weighted average: Σ (share_i × log_density_i)
        density_score = sum(
            (emp / emp_total) * cbsa_log_density[cbsa]
            for cbsa, emp in valid_metros.items()
        )

        # Top 5 metros by employment share
        sorted_metros = sorted(valid_metros.items(), key=lambda x: x[1], reverse=True)
        top_metros = []
        for cbsa, emp in sorted_metros[:5]:
            top_metros.append({
                "cbsa": cbsa,
                "name": metro_names.get(cbsa, "Unknown"),
                "employment": int(emp),
                "share": round(emp / emp_total, 4),
            })

        cbsa_count = len(valid_metros)
        scores[soc] = {
            "density_score": round(density_score, 6),
            "cbsa_count": cbsa_count,
            "low_data": cbsa_count < 10,
            "national_metro_employment": int(emp_total),
            "top_metros": top_metros,
        }

    return scores, unscored


def compute_ranks_and_percentiles(scores):
    """
    Add rank (1=highest) and percentile to each scored occupation.
    """
    socs = sorted(scores.keys())
    values = np.array([scores[soc]["density_score"] for soc in socs])

    N = len(values)
    rank_order = np.argsort(-values)
    ranks = np.empty(N, dtype=int)
    ranks[rank_order] = np.arange(1, N + 1)

    percentiles = (N - ranks) / (N - 1) * 100.0

    for i, soc in enumerate(socs):
        scores[soc]["rank"] = int(ranks[i])
        scores[soc]["percentile"] = round(float(percentiles[i]), 2)


def run_sanity_checks(scores, cbsa_info, cbsa_log_density, matched_codes):
    """Run comprehensive sanity checks on density scores."""
    N = len(scores)
    all_pass = True

    def check(name, condition, detail=""):
        nonlocal all_pass
        status = "PASS" if condition else "FAIL"
        if not condition:
            all_pass = False
        print(f"  [{status}] {name}" + (f" — {detail}" if detail else ""))

    print("\n" + "=" * 60)
    print("SANITY CHECKS")
    print("=" * 60)

    density_values = np.array([s["density_score"] for s in scores.values()])

    # 1. No NaN or Inf
    check("No NaN density scores", not np.any(np.isnan(density_values)))
    check("No Inf density scores", not np.any(np.isinf(density_values)))

    # 2. Meaningful spread
    check("Meaningful spread (std > 0.1)", density_values.std() > 0.1,
          f"std={density_values.std():.4f}")

    # 3. Score range is reasonable (log scale, expect roughly 2-8)
    check("Min density score > 1.0", density_values.min() > 1.0,
          f"min={density_values.min():.4f}")
    check("Max density score < 10.0", density_values.max() < 10.0,
          f"max={density_values.max():.4f}")

    # 4. CBSA count
    cbsa_count = len(cbsa_log_density)
    check("CBSA count in range 350-420", 350 <= cbsa_count <= 420,
          f"{cbsa_count} CBSAs")

    # 5. Scored occupation count
    check("Scored occupation count >= 680",
          N >= 680,
          f"{N} scored of {len(matched_codes)} matched")

    # 6. NYC density spot check
    print(f"\n  Metro density spot checks:")
    if "35620" in cbsa_log_density:
        nyc_ld = cbsa_log_density["35620"]
        sorted_cbsas = sorted(cbsa_log_density.items(), key=lambda x: x[1], reverse=True)
        nyc_rank = next(i + 1 for i, (c, _) in enumerate(sorted_cbsas) if c == "35620")
        check("NYC (35620) in top 10 densest CBSAs",
              nyc_rank <= 10,
              f"rank {nyc_rank}/{len(cbsa_log_density)}, log_density={nyc_ld:.4f}")
    else:
        print("  [SKIP] NYC (35620) not in CBSA data")

    # 7. Face-validity occupation checks
    print(f"\n  Face-validity occupation checks:")
    median_score = np.median(density_values)

    # Software Developers — concentrated in dense tech hubs
    if "15-1252" in scores:
        sd = scores["15-1252"]
        check("Software Developers above median density",
              sd["density_score"] > median_score,
              f"score={sd['density_score']:.4f}, median={median_score:.4f}")

    # Financial Analysts — concentrated in NYC, SF, Chicago
    if "13-2051" in scores:
        fa = scores["13-2051"]
        check("Financial Analysts above median density",
              fa["density_score"] > median_score,
              f"score={fa['density_score']:.4f}")

    # Logging Workers — rural occupation
    if "45-4022" in scores:
        lw = scores["45-4022"]
        check("Logging Workers below median density",
              lw["density_score"] < median_score,
              f"score={lw['density_score']:.4f}")
    else:
        # Try another rural occupation: Forest and Conservation Workers
        if "45-4011" in scores:
            fc = scores["45-4011"]
            check("Forest/Conservation Workers below median density",
                  fc["density_score"] < median_score,
                  f"score={fc['density_score']:.4f}")
        else:
            print("  [SKIP] No rural occupation found for below-median check")

    # 8. Distribution stats
    print(f"\n  Density score distribution:")
    print(f"    Min: {density_values.min():.4f}  Max: {density_values.max():.4f}")
    print(f"    Mean: {density_values.mean():.4f}  Median: {np.median(density_values):.4f}  Std: {density_values.std():.4f}")
    for pct in [5, 25, 50, 75, 95]:
        print(f"    {pct}th percentile: {np.percentile(density_values, pct):.4f}")

    print(f"\n  Overall: {'ALL CHECKS PASSED' if all_pass else 'SOME CHECKS FAILED'}")
    return all_pass


def main():
    # Ensure output directory exists
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Verify input exists
    if not TRANSFERABILITY_FILE.exists():
        print(f"ERROR: Input file not found: {TRANSFERABILITY_FILE}")
        print("Run scripts/04_compute_transferability.py first.")
        sys.exit(1)

    print("Step 5: Compute Geographic (Metro) Density Scores")
    print("=" * 60)

    # Load matched SOC codes from Step 4
    print("\nLoading matched SOC codes from Step 4...")
    with open(TRANSFERABILITY_FILE, "r", encoding="utf-8") as f:
        transfer_data = json.load(f)
    matched_codes = sorted(transfer_data["scores"].keys())
    titles = {soc: v["title"] for soc, v in transfer_data["scores"].items()}
    print(f"  {len(matched_codes)} scored occupations from Step 4")

    # Download data files
    print("\n" + "=" * 60)
    print("DOWNLOADING DATA")
    print("=" * 60)

    print("\n1. OEWS MSA flat file (May 2024)...")
    download_file(OEWS_URL, OEWS_ZIP, 1_000_000, "OEWS MSA flat file")
    extract_zip(OEWS_ZIP, OEWS_DIR)

    print("\n2. Census CBSA gazetteer (2024)...")
    download_file(GAZETTEER_URL, GAZETTEER_ZIP, 5_000, "Census CBSA gazetteer")
    extract_zip(GAZETTEER_ZIP, GAZETTEER_DIR)

    # Parse gazetteer
    print("\n" + "=" * 60)
    print("PARSING DATA")
    print("=" * 60)

    print("\nParsing Census CBSA gazetteer...")
    cbsa_areas = parse_gazetteer(GAZETTEER_DIR)
    print(f"  {len(cbsa_areas)} CBSAs with land area data")

    # Parse OEWS
    print("\nParsing OEWS MSA data...")
    occ_by_metro, metro_total_emp, metro_names = parse_oews_msa(OEWS_DIR)
    print(f"  {len(occ_by_metro)} occupations with metro-level data")
    print(f"  {len(metro_total_emp)} MSAs with total employment")

    # Compute CBSA log-density
    print("\n" + "=" * 60)
    print("COMPUTING DENSITY")
    print("=" * 60)

    print("\nComputing CBSA log-density (total_emp / land_area)...")
    cbsa_log_density, cbsa_info = compute_cbsa_log_density(metro_total_emp, cbsa_areas)
    print(f"  {len(cbsa_log_density)} CBSAs with valid log-density")

    # Show top/bottom CBSAs
    sorted_cbsas = sorted(cbsa_info.items(), key=lambda x: x[1]["log_density"], reverse=True)
    print(f"\n  Top 5 densest CBSAs:")
    for cbsa, info in sorted_cbsas[:5]:
        print(f"    {cbsa}: {info['name']}")
        print(f"      emp={info['total_employment']:,}  area={info['land_area_sq_mi']:.0f} sq mi  log_density={info['log_density']:.4f}")
    print(f"\n  Bottom 5 least dense CBSAs:")
    for cbsa, info in sorted_cbsas[-5:]:
        print(f"    {cbsa}: {info['name']}")
        print(f"      emp={info['total_employment']:,}  area={info['land_area_sq_mi']:.0f} sq mi  log_density={info['log_density']:.4f}")

    # Compute occupation density scores
    print(f"\nComputing density scores for {len(matched_codes)} occupations...")
    scores, unscored = compute_density_scores(
        occ_by_metro, cbsa_log_density, matched_codes, metro_names
    )
    print(f"  Scored: {len(scores)} occupations")
    print(f"  Unscored: {len(unscored)} occupations (no OEWS metro data)")

    if unscored:
        print(f"  First 10 unscored: {unscored[:10]}")

    # Add ranks and percentiles
    compute_ranks_and_percentiles(scores)

    # Add titles
    for soc in scores:
        scores[soc]["title"] = titles.get(soc, soc)

    # Sanity checks
    all_pass = run_sanity_checks(scores, cbsa_info, cbsa_log_density, matched_codes)

    if not all_pass:
        print("\nWARNING: Some sanity checks failed. Review output carefully.")

    # Build output
    print("\n" + "=" * 60)
    print("WRITING OUTPUT")
    print("=" * 60)

    density_values = np.array([s["density_score"] for s in scores.values()])

    output = {
        "metadata": {
            "description": "Geographic metro density scores per occupation",
            "formula": "E[log(density)|occ] = sum_i(emp_occ_i / emp_occ_total) * log(emp_all_i / area_i)",
            "formula_source": "Manning & Aguirre (2026), NBER Working Paper 34705",
            "inputs": {
                "oews_msa": "BLS OEWS May 2024 MSA estimates (oesm24ma.zip)",
                "gazetteer": "Census 2024 CBSA Gazetteer (2024_Gaz_cbsa_national.zip)",
            },
            "employment_source_note": (
                "OEWS used for both per-occupation and total metro employment "
                "(nonfarm wage-and-salary workers). The formula uses employment "
                "shares (ratios), so this coverage is methodologically consistent."
            ),
            "low_data_flag": (
                "Occupations with low_data=true have fewer than 10 metros "
                "in OEWS data. Their density scores are valid but noisy — "
                "driven by whichever few metros report data. The front-end "
                "can use this flag to add caveats or offer a toggle."
            ),
            "total_cbsas": len(cbsa_log_density),
            "total_scored_occupations": len(scores),
            "low_data_occupations": sum(1 for s in scores.values() if s["low_data"]),
            "unscored_occupations": len(unscored),
            "score_range": {
                "min": round(float(density_values.min()), 6),
                "max": round(float(density_values.max()), 6),
                "mean": round(float(density_values.mean()), 6),
                "std": round(float(density_values.std()), 6),
            },
            "generated_by": "scripts/05_compute_density.py",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "cbsa_lookup": {},
        "occupation_density": {},
    }

    # CBSA lookup table (sorted by log_density descending for readability)
    for cbsa, info in sorted(cbsa_info.items(), key=lambda x: x[1]["log_density"], reverse=True):
        output["cbsa_lookup"][cbsa] = info

    # Occupation density scores (sorted by SOC code)
    for soc in sorted(scores.keys()):
        output["occupation_density"][soc] = scores[soc]

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)
    file_size = OUTPUT_FILE.stat().st_size / (1024 * 1024)
    print(f"\n  Wrote {OUTPUT_FILE} ({file_size:.2f} MB)")

    # Sample output
    print("\n" + "=" * 60)
    print("SAMPLE OUTPUT")
    print("=" * 60)

    # Find notable occupations
    N = len(scores)
    scored_socs = sorted(scores.keys(), key=lambda s: scores[s]["rank"])
    top_soc = scored_socs[0]
    bottom_soc = scored_socs[-1]
    median_soc = scored_socs[N // 2]

    notable = [
        ("Highest density", top_soc),
        ("Lowest density", bottom_soc),
        ("Median density", median_soc),
    ]

    for label, soc in [("Software Developers", "15-1252"), ("Registered Nurses", "29-1141")]:
        if soc in scores:
            notable.append((label, soc))

    for label, soc in notable:
        s = scores[soc]
        print(f"\n  {label}:")
        print(f"    {soc}: {s['title']}")
        print(f"    Density score = {s['density_score']:.4f} (rank {s['rank']}/{N}, {s['percentile']:.1f}th percentile)")
        print(f"    Present in {s['cbsa_count']} metros, national metro employment: {s['national_metro_employment']:,}")
        if s["top_metros"]:
            print(f"    Top-3 metros:")
            for m in s["top_metros"][:3]:
                print(f"      {m['share']:.1%}  {m['cbsa']} — {m['name']} ({m['employment']:,} workers)")


if __name__ == "__main__":
    main()
